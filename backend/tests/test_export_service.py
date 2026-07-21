"""Unit tests for app.services.export_service.

Tests the four export functions by mocking the DB session and SQLAlchemy
ORM functions so they run fully offline.
"""

from __future__ import annotations

from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from tests.conftest import _patch_sa


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_mock_project(**kwargs) -> MagicMock:
    p = MagicMock()
    p.id = kwargs.get("id", 1)
    p.name = kwargs.get("name", "Test Office")
    p.client = kwargs.get("client", "Acme Corp")
    p.project_code = kwargs.get("project_code", "PROJ-001")
    p.location = kwargs.get("location", "Bangalore")
    p.status = kwargs.get("status", "active")
    p.currency = kwargs.get("currency", "INR")
    p.description = kwargs.get("description", "Office fit-out")
    p.start_date = None
    p.end_date = None
    p.is_deleted = False
    return p


def _make_mock_boq_item(**kwargs) -> MagicMock:
    item = MagicMock()
    item.id = kwargs.get("id", 101)
    item.project_id = kwargs.get("project_id", 1)
    item.description = kwargs.get("description", "Tile flooring")
    item.quantity = kwargs.get("quantity", 100.0)
    item.unit = kwargs.get("unit", "sqm")
    item.rate = kwargs.get("rate", 850.0)
    item.wastage_pct = kwargs.get("wastage_pct", 10.0)
    item.labour_rate = kwargs.get("labour_rate", 800.0)
    item.transport_rate = kwargs.get("transport_rate", 0.0)
    item.transport_pct = kwargs.get("transport_pct", 0.0)
    item.overhead_pct = kwargs.get("overhead_pct", 10.0)
    item.margin_pct = kwargs.get("margin_pct", 15.0)
    item.discount_pct = kwargs.get("discount_pct", 10.0)
    item.gst_rate = kwargs.get("gst_rate", 18.0)
    item.total = kwargs.get("total", 85000.0)
    item.category = kwargs.get("category", "Flooring")
    item.material_name = kwargs.get("material_name", "Tile 600x600")
    item.material_id = kwargs.get("material_id", 42)
    item.vendor_id = kwargs.get("vendor_id", 7)
    item.object_id = kwargs.get("object_id", None)
    item.sort_order = kwargs.get("sort_order", 1)
    item.is_deleted = False
    item.detected_object = None
    return item


def _make_mock_version(**kwargs) -> MagicMock:
    cv = MagicMock()
    cv.id = kwargs.get("id", 1)
    cv.name = kwargs.get("name", "v1.0")
    cv.version_number = kwargs.get("version_number", 1)
    cv.status = kwargs.get("status", "draft")
    cv.is_deleted = False
    return cv


def _configure_db_for_boq(db, project=None, boq_items=None, version=None):
    """Configure a mock DB session to return the given project/items/version
    across three consecutive execute calls (project, items, version)."""
    project_result = MagicMock()
    project_result.scalar_one_or_none.return_value = project
    items_result = MagicMock()
    items_result.unique.return_value.scalars.return_value.all.return_value = boq_items or []
    version_result = MagicMock()
    version_result.scalar_one_or_none.return_value = version
    db.execute.side_effect = [project_result, items_result, version_result]


# =========================================================================
# _boq_items_to_dicts (no SA dependency)
# =========================================================================


class TestBoqItemsToDicts:
    def _import_and_test(self):
        import app.services.export_service as svc
        return svc

    def test_empty(self):
        svc = self._import_and_test()
        with _patch_sa(svc):
            assert svc._boq_items_to_dicts([]) == ([], [])

    def test_single_item(self):
        svc = self._import_and_test()
        item = _make_mock_boq_item()
        with _patch_sa(svc):
            dicts, breakdowns = svc._boq_items_to_dicts([item])
        assert len(dicts) == 1
        assert len(breakdowns) == 1
        assert dicts[0]["id"] == 101
        assert dicts[0]["description"] == "Tile flooring"

    def test_trade_from_category(self):
        svc = self._import_and_test()
        item = _make_mock_boq_item(category="Flooring")
        item.detected_object = None
        with _patch_sa(svc):
            dicts, _ = svc._boq_items_to_dicts([item])
        assert dicts[0]["trade"] == "Flooring"

    def test_trade_from_detected_object(self):
        svc = self._import_and_test()
        item = _make_mock_boq_item(category=None)
        obj = MagicMock()
        obj.object_type = "Wall"
        item.detected_object = obj
        with _patch_sa(svc):
            dicts, _ = svc._boq_items_to_dicts([item])
        assert dicts[0]["trade"] == "Wall"


# =========================================================================
# export_boq_xlsx
# =========================================================================


class TestExportBoqXlsx:
    @pytest.mark.asyncio
    async def test_generates_valid_xlsx(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(),
            boq_items=[_make_mock_boq_item()],
            version=_make_mock_version(),
        )

        with _patch_sa(svc):
            data, filename = await svc.export_boq_xlsx(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0
        assert filename.endswith(".xlsx")
        # Verify workbook content
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(data))
        assert "BOQ Summary" in wb.sheetnames
        assert "BOQ by Trade" in wb.sheetnames
        assert "Cost Breakdown" in wb.sheetnames
        assert "Material Summary" in wb.sheetnames
        ws = wb["BOQ Summary"]
        assert ws["A1"].value is not None

    @pytest.mark.asyncio
    async def test_empty_boq_items(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(),
            boq_items=[],
            version=None,
        )

        with _patch_sa(svc):
            data, filename = await svc.export_boq_xlsx(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(data))
        assert "BOQ Summary" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_project_not_found(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(db, project=None)

        with _patch_sa(svc):
            with pytest.raises(ValueError, match="Project.*not found"):
                await svc.export_boq_xlsx(999, db)

    @pytest.mark.asyncio
    async def test_filename_safe(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(name="Test / Project"),
            boq_items=[_make_mock_boq_item()],
            version=None,
        )

        with _patch_sa(svc):
            _, filename = await svc.export_boq_xlsx(1, db)
        assert "/" not in filename
        assert " " not in filename


# =========================================================================
# export_proposal_pdf
# =========================================================================


class TestExportProposalPdf:
    @pytest.mark.asyncio
    async def test_generates_valid_pdf(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(),
            boq_items=[_make_mock_boq_item()],
            version=_make_mock_version(),
        )

        with _patch_sa(svc):
            data, filename = await svc.export_proposal_pdf(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0
        assert data.startswith(b"%PDF")
        assert filename.endswith(".pdf")

    @pytest.mark.asyncio
    async def test_empty_data(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(name="Empty Proposal"),
            boq_items=[],
            version=None,
        )

        with _patch_sa(svc):
            data, filename = await svc.export_proposal_pdf(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0
        assert data.startswith(b"%PDF")

    @pytest.mark.asyncio
    async def test_project_not_found(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(db, project=None)

        with _patch_sa(svc):
            with pytest.raises(ValueError, match="Project.*not found"):
                await svc.export_proposal_pdf(999, db)


# =========================================================================
# export_purchase_list
# =========================================================================


class TestExportPurchaseList:
    @pytest.mark.asyncio
    async def test_generates_valid_xlsx(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(),
            boq_items=[_make_mock_boq_item()],
            version=None,
        )

        with _patch_sa(svc):
            data, filename = await svc.export_purchase_list(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0
        assert filename.endswith(".xlsx")
        from openpyxl import load_workbook
        from io import BytesIO
        wb = load_workbook(BytesIO(data))
        assert "Purchase by Vendor" in wb.sheetnames
        assert "By Urgency" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_empty_items(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(),
            boq_items=[],
            version=None,
        )

        with _patch_sa(svc):
            data, filename = await svc.export_purchase_list(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0

    @pytest.mark.asyncio
    async def test_project_not_found(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(db, project=None)

        with _patch_sa(svc):
            with pytest.raises(ValueError, match="Project.*not found"):
                await svc.export_purchase_list(999, db)


# =========================================================================
# export_client_presentation
# =========================================================================


class TestExportClientPresentation:
    @pytest.mark.asyncio
    async def test_generates_valid_pdf(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(),
            boq_items=[_make_mock_boq_item()],
            version=_make_mock_version(),
        )

        with _patch_sa(svc):
            data, filename = await svc.export_client_presentation(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0
        assert data.startswith(b"%PDF")
        assert filename.endswith(".pdf")

    @pytest.mark.asyncio
    async def test_empty_data(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(
            db,
            project=_make_mock_project(name="Empty"),
            boq_items=[],
            version=None,
        )

        with _patch_sa(svc):
            data, filename = await svc.export_client_presentation(1, db)
        assert isinstance(data, bytes)
        assert len(data) > 0
        assert data.startswith(b"%PDF")

    @pytest.mark.asyncio
    async def test_project_not_found(self):
        import app.services.export_service as svc

        db = AsyncMock()
        _configure_db_for_boq(db, project=None)

        with _patch_sa(svc):
            with pytest.raises(ValueError, match="Project.*not found"):
                await svc.export_client_presentation(999, db)
