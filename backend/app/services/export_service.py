"""
Export Service — BOQ Excel, Proposal PDF, Purchase List, Client Presentation.

Each function:
1. Loads project + BOQ data via the provided async DB session
2. Generates the document in memory (BytesIO)
3. Returns (bytes, filename_with_extension)

Large exports stream through BytesIO — no temp files on disk.
"""
from __future__ import annotations

import io
import logging
import math
from datetime import datetime, timedelta
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
    HRFlowable,
    Image,
    Flowable,
)
from sqlalchemy import select
from sqlalchemy.orm import joinedload

from app.models.core import Project
from app.models.detection import BOQItem, CostVersion, DetectedObject
from app.models.reference import Material, Vendor
from app.services.cost_engine import compute_line_item, CostBreakdown, _aggregate_python

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
CURRENCY_SYMBOL = "₹"
INR_FORMAT = '#,##0.00'
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=10)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", bold=True, size=10)

PROPOSAL_BLUE = HexColor("#1F4E79")
PROPOSAL_ACCENT = HexColor("#D4E6F1")


def _r2(value: float) -> float:
    """Round to 2 decimal places."""
    return round(value, 2)


# ======================================================================
# Helper: load data
# ======================================================================


async def _load_project_with_boq(project_id: int, db) -> tuple[Project, list[BOQItem], CostVersion | None]:
    """Load project, BOQ items (with detected objects), and latest cost version."""
    result = await db.execute(
        select(Project).where(
            Project.id == project_id,
            Project.is_deleted == False,  # noqa: E712
        )
    )
    project = result.scalar_one_or_none()
    if project is None:
        raise ValueError(f"Project {project_id} not found")

    items_result = await db.execute(
        select(BOQItem)
        .options(joinedload(BOQItem.detected_object))
        .where(
            BOQItem.project_id == project_id,
            BOQItem.is_deleted == False,  # noqa: E712
        )
        .order_by(BOQItem.sort_order, BOQItem.id)
    )
    boq_items = list(items_result.unique().scalars().all())

    version_result = await db.execute(
        select(CostVersion)
        .where(
            CostVersion.project_id == project_id,
            CostVersion.is_deleted == False,  # noqa: E712
        )
        .order_by(CostVersion.version_number.desc())
        .limit(1)
    )
    latest_version = version_result.scalar_one_or_none()

    return project, boq_items, latest_version


def _boq_items_to_dicts(
    boq_items: list[BOQItem],
) -> tuple[list[dict], list[CostBreakdown]]:
    """Convert BOQItem ORM rows to item dicts and compute cost breakdowns."""
    item_dicts: list[dict] = []
    for item in boq_items:
        obj = item.detected_object
        trade = item.category or (obj.object_type if obj else "Other")
        item_dicts.append(
            {
                "id": item.id,
                "item_id": item.id,
                "description": item.description,
                "quantity": item.quantity,
                "unit": item.unit,
                "rate": item.rate,
                "wastage_pct": item.wastage_pct,
                "labour_rate": item.labour_rate,
                "transport_rate": item.transport_rate,
                "transport_pct": item.transport_pct,
                "overhead_pct": item.overhead_pct,
                "margin_pct": item.margin_pct,
                "discount_pct": item.discount_pct,
                "gst_rate": item.gst_rate,
                "trade": trade,
                "category": item.category,
                "material_name": item.material_name,
                "material_id": item.material_id,
                "vendor_id": item.vendor_id,
            }
        )
    breakdowns = [compute_line_item(it) for it in item_dicts]
    return item_dicts, breakdowns


# ======================================================================
# 1. BOQ Excel (openpyxl)
# ======================================================================


async def export_boq_xlsx(project_id: int, db) -> tuple[bytes, str]:
    """Generate a multi-sheet BOQ Excel workbook.

    Sheets:
        - BOQ Summary: project metadata + grand total
        - BOQ by Trade: grouped line items
        - Cost Breakdown: per-item cost details
        - Material Summary: material catalogue used
    """
    project, boq_items, latest_version = await _load_project_with_boq(project_id, db)
    item_dicts, breakdowns = _boq_items_to_dicts(boq_items)
    totals = _aggregate_python(breakdowns) if breakdowns else {}

    wb = Workbook()

    _write_boq_summary_sheet(wb, project, boq_items, latest_version, totals)
    _write_boq_by_trade_sheet(wb, project, boq_items, item_dicts, breakdowns)
    _write_cost_breakdown_sheet(wb, breakdowns, item_dicts)
    _write_material_summary_sheet(wb, item_dicts)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = project.name.replace(" ", "_").replace("/", "_")[:50] or f"project_{project_id}"
    filename = f"BOQ_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename


def _style_header_row(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_body_cell(cell, is_currency: bool = False):
    """Apply body styling to a cell."""
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    if is_currency:
        cell.number_format = INR_FORMAT


def _auto_width(ws, min_width: int = 8, max_width: int = 50):
    """Auto-adjust column widths."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            # Rough char-width calculation
            lengths.append(min(len(val), max_width))
        best = max(lengths) + 2 if lengths else min_width
        ws.column_dimensions[col_letter].width = max(min_width, min(best, max_width))


def _write_boq_summary_sheet(
    wb: Workbook,
    project: Project,
    boq_items: list[BOQItem],
    latest_version: CostVersion | None,
    totals: dict[str, Any],
):
    """Sheet 1: BOQ Summary."""
    ws = wb.active
    ws.title = "BOQ Summary"
    ws.sheet_properties.tabColor = "1F4E79"

    # Title
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = f"Bill of Quantities — {project.name}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Project info
    row = 4
    info_data = [
        ("Project Name", project.name),
        ("Client", project.client or "—"),
        ("Project Code", project.project_code or "—"),
        ("Location", project.location or "—"),
        ("Status", project.status),
        ("Currency", project.currency),
        ("Version", latest_version.name if latest_version else "N/A"),
        ("Version Status", latest_version.status if latest_version else "—"),
        ("Total Items", str(len(boq_items))),
        ("Grand Total", f'{CURRENCY_SYMBOL}{totals.get("grand_total", 0):,.2f}'),
    ]
    for i, (label, value) in enumerate(info_data):
        r = row + i
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=r, column=1).fill = SUBHEADER_FILL
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        val_cell = ws.cell(row=r, column=2, value=value)
        val_cell.font = BODY_FONT
        val_cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def _write_boq_by_trade_sheet(
    wb: Workbook,
    project: Project,
    boq_items: list[BOQItem],
    item_dicts: list[dict],
    breakdowns: list[CostBreakdown],
):
    """Sheet 2: BOQ by Trade — grouped table."""
    ws = wb.create_sheet("BOQ by Trade")
    ws.sheet_properties.tabColor = "2E75B6"

    headers = ["Trade", "Description", "Qty", "Unit", "Rate", "Total"]
    col_widths = [18, 50, 12, 8, 14, 16]

    # Group by trade
    trade_groups: dict[str, list[tuple[dict, CostBreakdown]]] = {}
    for item_dict, bd in zip(item_dicts, breakdowns):
        trade = str(item_dict.get("trade", item_dict.get("category", "Other")))
        trade_groups.setdefault(trade, []).append((item_dict, bd))

    current_row = 1

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=f"BOQ by Trade — {project.name}")
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center")
    current_row = 3

    for trade_name in sorted(trade_groups.keys()):
        items = trade_groups[trade_name]

        # Trade header
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(headers),
        )
        trade_cell = ws.cell(row=current_row, column=1, value=trade_name)
        trade_cell.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        trade_cell.fill = SUBHEADER_FILL
        trade_cell.alignment = Alignment(horizontal="left")
        for c in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
            ws.cell(row=current_row, column=c).fill = SUBHEADER_FILL
        current_row += 1

        # Column headers
        for ci, h in enumerate(headers, 1):
            ws.cell(row=current_row, column=ci, value=h)
        _style_header_row(ws, current_row, len(headers))
        current_row += 1

        # Items
        trade_total = 0.0
        for item_dict, bd in items:
            total_val = _r2(bd.grand_total)
            trade_total += total_val
            row_data = [
                "",
                item_dict.get("description", ""),
                item_dict.get("quantity", 0),
                item_dict.get("unit", "nos"),
                _r2(item_dict.get("rate", 0)),
                total_val,
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=ci, value=val)
                _style_body_cell(cell, is_currency=(ci in (5, 6)))
            current_row += 1

        # Trade subtotal
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=4,
        )
        sub_cell = ws.cell(row=current_row, column=1, value=f"Subtotal — {trade_name}")
        sub_cell.font = TOTAL_FONT
        sub_cell.fill = TOTAL_FILL
        sub_cell.border = THIN_BORDER
        sub_cell.alignment = Alignment(horizontal="right")
        for c in range(2, 5):
            ws.cell(row=current_row, column=c).fill = TOTAL_FILL
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        total_cell = ws.cell(row=current_row, column=5, value=_r2(trade_total))
        total_cell.font = TOTAL_FONT
        total_cell.fill = TOTAL_FILL
        total_cell.border = THIN_BORDER
        total_cell.number_format = INR_FORMAT
        ws.cell(row=current_row, column=6).fill = TOTAL_FILL
        ws.cell(row=current_row, column=6).border = THIN_BORDER
        current_row += 2  # blank row between trades

    # Grand total
    current_row += 1
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    gt_label = ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    gt_label.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    gt_label.fill = HEADER_FILL
    gt_label.alignment = Alignment(horizontal="right")
    for c in range(2, 5):
        ws.cell(row=current_row, column=c).fill = HEADER_FILL
        ws.cell(row=current_row, column=c).border = THIN_BORDER
    gt_val = sum(
        _r2(bd.grand_total) for bd in breakdowns
    ) if breakdowns else 0
    gt_cell = ws.cell(row=current_row, column=5, value=_r2(gt_val))
    gt_cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    gt_cell.fill = HEADER_FILL
    gt_cell.border = THIN_BORDER
    gt_cell.number_format = INR_FORMAT
    ws.cell(row=current_row, column=6).fill = HEADER_FILL
    ws.cell(row=current_row, column=6).border = THIN_BORDER

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_cost_breakdown_sheet(
    wb: Workbook,
    breakdowns: list[CostBreakdown],
    item_dicts: list[dict],
):
    """Sheet 3: Per-item cost breakdown."""
    ws = wb.create_sheet("Cost Breakdown")
    ws.sheet_properties.tabColor = "548235"

    headers = [
        "#", "Description", "Qty", "Rate",
        "Material", "Labour", "Transport", "Overhead",
        "Margin", "GST", "Grand Total",
    ]
    col_widths = [5, 40, 10, 10, 14, 12, 12, 12, 10, 10, 14]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].value = "Cost Breakdown — Detailed View"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="548235")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    _style_header_row(ws, 3, len(headers))

    grand_total = 0.0
    for i, (bd, item_dict) in enumerate(zip(breakdowns, item_dicts), 1):
        row = 3 + i
        row_data = [
            i,
            item_dict.get("description", ""),
            item_dict.get("quantity", 0),
            _r2(item_dict.get("rate", 0)),
            _r2(bd.material_cost),
            _r2(bd.labour_cost),
            _r2(bd.transport_cost),
            _r2(bd.overhead_cost),
            _r2(bd.margin_cost),
            _r2(bd.gst_amount),
            _r2(bd.grand_total),
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            is_curr = ci >= 4
            _style_body_cell(cell, is_currency=is_curr)
        grand_total += bd.grand_total

    # Grand total
    total_row = 3 + len(breakdowns) + 1
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
    gt_label = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    gt_label.font = Font(name="Calibri", bold=True, size=11, color="548235")
    gt_label.fill = TOTAL_FILL
    gt_label.alignment = Alignment(horizontal="right")
    for c in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=c).fill = TOTAL_FILL
        ws.cell(row=total_row, column=c).border = THIN_BORDER
        ws.cell(row=total_row, column=c).font = TOTAL_FONT
    gt_cell = ws.cell(row=total_row, column=11, value=_r2(grand_total))
    gt_cell.number_format = INR_FORMAT

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_material_summary_sheet(
    wb: Workbook,
    item_dicts: list[dict],
):
    """Sheet 4: Material/Item summary."""
    ws = wb.create_sheet("Material Summary")
    ws.sheet_properties.tabColor = "BF8F00"

    headers = ["Description", "Material Name", "Vendor ID", "Qty", "Unit", "Rate", "Total"]
    col_widths = [40, 25, 12, 12, 8, 14, 16]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].value = "Material / Item Summary"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="BF8F00")
    ws["A1"].alignment = Alignment(horizontal="center")

    for ci, h in enumerate(headers, 1):
        ws.cell(row=3, column=ci, value=h)
    _style_header_row(ws, 3, len(headers))

    for i, item_dict in enumerate(item_dicts, 1):
        row = 3 + i
        total = _r2(item_dict.get("quantity", 0) * item_dict.get("rate", 0))
        row_data = [
            item_dict.get("description", ""),
            item_dict.get("material_name", "—"),
            item_dict.get("vendor_id", "—"),
            item_dict.get("quantity", 0),
            item_dict.get("unit", "nos"),
            _r2(item_dict.get("rate", 0)),
            total,
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            _style_body_cell(cell, is_currency=(ci in (6, 7)))

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ======================================================================
# 2. Proposal PDF (reportlab)
# ======================================================================


async def export_proposal_pdf(project_id: int, db) -> tuple[bytes, str]:
    """Generate a formal proposal PDF with cover page, BOQ, cost summary, etc."""
    project, boq_items, latest_version = await _load_project_with_boq(project_id, db)
    item_dicts, breakdowns = _boq_items_to_dicts(boq_items)
    totals = _aggregate_python(breakdowns) if breakdowns else {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        title=f"Proposal — {project.name}",
        author="Auto Cost Engine",
    )

    styles = getSampleStyleSheet()
    _add_proposal_styles(styles)

    elements: list[Flowable] = []

    # ── Page 1: Cover ────────────────────────────────────────────────
    _add_cover_page(elements, styles, project, latest_version)

    # ── Page 2: Executive Summary ────────────────────────────────────
    elements.append(PageBreak())
    _add_executive_summary(elements, styles, project, boq_items, totals, latest_version)

    # ── Page 3: Detected Objects Summary ─────────────────────────────
    if boq_items:
        elements.append(PageBreak())
        _add_detected_objects_summary(elements, styles, boq_items, item_dicts)

    # ── Pages 4-5: BOQ by Trade ──────────────────────────────────────
    if breakdowns:
        elements.append(PageBreak())
        _add_boq_table_pdf(elements, styles, item_dicts, breakdowns)

    # ── Page 6: Cost Summary (pie chart placeholder) ─────────────────
    elements.append(PageBreak())
    _add_cost_summary_pdf(elements, styles, totals)

    # ── Page 7: Selected Materials Highlights ────────────────────────
    elements.append(PageBreak())
    _add_materials_highlights_pdf(elements, styles, item_dicts)

    # ── Page 8: Terms & Conditions ───────────────────────────────────
    elements.append(PageBreak())
    _add_terms_and_conditions(elements, styles)

    # Build with page numbers
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(
            A4[0] / 2, 12 * mm,
            f"Page {doc.page} | {project.name} | {datetime.now().strftime('%d-%b-%Y')}",
        )
        # Header line
        canvas.setStrokeColor(HexColor("#1F4E79"))
        canvas.setLineWidth(0.5)
        canvas.line(20 * mm, A4[1] - 15 * mm, A4[0] - 20 * mm, A4[1] - 15 * mm)
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

    safe_name = project.name.replace(" ", "_").replace("/", "_")[:50] or f"project_{project_id}"
    filename = f"Proposal_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buf.getvalue(), filename


def _add_proposal_styles(styles):
    """Add custom paragraph styles for the proposal."""
    styles.add(
        ParagraphStyle(
            name="CoverTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=28,
            textColor=PROPOSAL_BLUE,
            spaceAfter=12 * mm,
            alignment=TA_CENTER,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CoverSub",
            parent=styles["Normal"],
            fontSize=14,
            textColor=colors.HexColor("#555555"),
            spaceAfter=6 * mm,
            alignment=TA_CENTER,
        )
    )
    styles.add(
        ParagraphStyle(
            name="SectionHeading",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            textColor=PROPOSAL_BLUE,
            spaceBefore=6 * mm,
            spaceAfter=4 * mm,
        )
    )
    styles.add(
        ParagraphStyle(
            name="InfoLabel",
            parent=styles["Normal"],
            fontSize=10,
            textColor=colors.HexColor("#666666"),
            spaceBefore=1 * mm,
            spaceAfter=1 * mm,
        )
    )
    styles.add(
        ParagraphStyle(
            name="InfoValue",
            parent=styles["Normal"],
            fontSize=11,
            textColor=colors.black,
            spaceBefore=0,
            spaceAfter=2 * mm,
        )
    )
    styles.add(
        ParagraphStyle(
            name="BodySmall",
            parent=styles["Normal"],
            fontSize=9,
            leading=12,
        )
    )
    styles.add(
        ParagraphStyle(
            name="TermsBody",
            parent=styles["Normal"],
            fontSize=9,
            leading=14,
            spaceAfter=3 * mm,
        )
    )


def _add_cover_page(elements, styles, project, latest_version):
    """Cover page with project name, client, date, proposal number."""
    elements.append(Spacer(1, 40 * mm))

    # Company logo placeholder
    elements.append(Paragraph(
        "<b>AUTO COST ENGINE</b>",
        ParagraphStyle(
            "LogoStyle",
            fontName="Helvetica-Bold",
            fontSize=18,
            textColor=PROPOSAL_BLUE,
            alignment=TA_CENTER,
        ),
    ))
    elements.append(Paragraph(
        "Estimation & Costing Solutions",
        ParagraphStyle(
            "LogoTag",
            fontName="Helvetica",
            fontSize=10,
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER,
        ),
    ))
    elements.append(Spacer(1, 20 * mm))

    # Horizontal rule
    elements.append(HRFlowable(
        width="80%", thickness=2,
        color=PROPOSAL_BLUE, spaceBefore=0, spaceAfter=10 * mm,
    ))

    elements.append(Paragraph("PROPOSAL", styles["CoverTitle"]))
    elements.append(Paragraph(
        f"<b>{project.name}</b>", styles["CoverSub"],
    ))
    if project.client:
        elements.append(Paragraph(
            f"Prepared for: {project.client}", styles["CoverSub"],
        ))

    elements.append(Spacer(1, 15 * mm))
    elements.append(Paragraph(
        f"Proposal No: PROP-{datetime.now().strftime('%Y%m%d-%H%M%S')}",
        styles["InfoLabel"],
    ))
    elements.append(Paragraph(
        f"Date: {datetime.now().strftime('%d %B %Y')}",
        styles["InfoLabel"],
    ))
    if latest_version:
        elements.append(Paragraph(
            f"Version: {latest_version.name or f'v{latest_version.version_number}'}",
            styles["InfoLabel"],
        ))
    if project.location:
        elements.append(Paragraph(
            f"Location: {project.location}", styles["InfoLabel"],
        ))

    elements.append(Spacer(1, 30 * mm))
    elements.append(HRFlowable(
        width="60%", thickness=0.5,
        color=colors.grey, spaceBefore=0, spaceAfter=6 * mm,
    ))
    elements.append(Paragraph(
        "CONFIDENTIAL — This document contains proprietary information.",
        ParagraphStyle(
            "Confidential",
            fontName="Helvetica-Oblique",
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER,
        ),
    ))


def _add_executive_summary(elements, styles, project, boq_items, totals, latest_version):
    """Page 2: Executive summary with key numbers."""
    elements.append(Paragraph("Executive Summary", styles["SectionHeading"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=PROPOSAL_BLUE,
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    # Scope
    elements.append(Paragraph(
        f"This proposal presents a detailed cost estimate for the <b>{project.name}</b> "
        f"project{f' located at {project.location}' if project.location else ''}.",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 4 * mm))

    if project.description:
        elements.append(Paragraph(project.description, styles["Normal"]))
        elements.append(Spacer(1, 4 * mm))

    # Key numbers table
    elements.append(Paragraph("<b>Key Financial Summary</b>", styles["Normal"]))
    elements.append(Spacer(1, 3 * mm))

    total_area = sum(
        (item.get("length", 0) or 0) * (item.get("width", 0) or 0)
        for item in [{}]  # simplified — real area from detected objects
    )

    summary_data = [
        ["Metric", "Value"],
        ["Total Items", str(len(boq_items))],
        ["Total Materials", f'{CURRENCY_SYMBOL}{totals.get("total_materials", 0):,.2f}'],
        ["Total Labour", f'{CURRENCY_SYMBOL}{totals.get("total_labour", 0):,.2f}'],
        ["Total Transport", f'{CURRENCY_SYMBOL}{totals.get("total_transport", 0):,.2f}'],
        ["Total Overhead", f'{CURRENCY_SYMBOL}{totals.get("total_overhead", 0):,.2f}'],
        ["Total Margin", f'{CURRENCY_SYMBOL}{totals.get("total_margin", 0):,.2f}'],
        ["GST", f'{CURRENCY_SYMBOL}{totals.get("total_gst", 0):,.2f}'],
        ["", ""],
        ["<b>GRAND TOTAL</b>", f'<b>{CURRENCY_SYMBOL}{totals.get("grand_total", 0):,.2f}</b>'],
    ]

    table = Table(summary_data, colWidths=[120 * mm, 60 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), PROPOSAL_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (0, -1), (-1, -1), HexColor("#E2EFDA")),
    ]))
    elements.append(table)

    if project.start_date and project.end_date:
        elements.append(Spacer(1, 4 * mm))
        duration = (project.end_date - project.start_date).days
        elements.append(Paragraph(
            f"<b>Project Duration:</b> {project.start_date.strftime('%d-%b-%Y')} to "
            f"{project.end_date.strftime('%d-%b-%Y')} ({duration} days)",
            styles["Normal"],
        ))

    if latest_version:
        elements.append(Spacer(1, 3 * mm))
        elements.append(Paragraph(
            f"<b>Version:</b> {latest_version.name or f'v{latest_version.version_number}'} "
            f"(Status: {latest_version.status})",
            styles["Normal"],
        ))


def _add_detected_objects_summary(elements, styles, boq_items, item_dicts):
    """Page 3: Detected objects types and counts."""
    elements.append(Paragraph("Detected Objects Summary", styles["SectionHeading"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=PROPOSAL_BLUE,
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    # Count by trade/category
    trade_counts: dict[str, int] = {}
    for item_dict in item_dicts:
        trade = str(item_dict.get("trade", item_dict.get("category", "Other")))
        trade_counts[trade] = trade_counts.get(trade, 0) + 1

    obj_data = [["Trade / Category", "Item Count"]]
    for trade, count in sorted(trade_counts.items()):
        obj_data.append([trade, str(count)])

    obj_data.append(["", ""])
    obj_data.append(["<b>Total</b>", f"<b>{sum(trade_counts.values())}</b>"])

    table = Table(obj_data, colWidths=[130 * mm, 50 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), PROPOSAL_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (0, -1), (-1, -1), HexColor("#E2EFDA")),
    ]))
    elements.append(table)


def _add_boq_table_pdf(elements, styles, item_dicts, breakdowns):
    """Pages 4-5: Full BOQ table grouped by trade."""
    elements.append(Paragraph("Bill of Quantities — Detailed Breakdown", styles["SectionHeading"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=PROPOSAL_BLUE,
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    # Group by trade
    trade_groups: dict[str, list] = {}
    for item_dict, bd in zip(item_dicts, breakdowns):
        trade = str(item_dict.get("trade", item_dict.get("category", "Other")))
        trade_groups.setdefault(trade, []).append((item_dict, bd))

    for trade_name in sorted(trade_groups.keys()):
        items = trade_groups[trade_name]

        elements.append(Paragraph(f"<b>{trade_name}</b>", styles["Normal"]))
        elements.append(Spacer(1, 2 * mm))

        table_data = [["#", "Description", "Qty", "Unit", "Rate", "Total"]]
        trade_total = 0.0
        for i, (item_dict, bd) in enumerate(items, 1):
            total_val = _r2(bd.grand_total)
            trade_total += total_val
            table_data.append([
                str(i),
                item_dict.get("description", "")[:60],
                f'{item_dict.get("quantity", 0):.2f}',
                item_dict.get("unit", "nos"),
                f'{CURRENCY_SYMBOL}{item_dict.get("rate", 0):,.2f}',
                f'{CURRENCY_SYMBOL}{total_val:,.2f}',
            ])

        table_data.append(["", "", "", "", "", ""])
        table_data.append([
            "", f"<b>Subtotal — {trade_name}</b>", "", "", "",
            f"<b>{CURRENCY_SYMBOL}{_r2(trade_total):,.2f}</b>",
        ])

        col_widths = [12 * mm, 78 * mm, 16 * mm, 12 * mm, 22 * mm, 24 * mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -2), 0.3, colors.lightgrey),
            ("BACKGROUND", (0, 0), (-1, 0), PROPOSAL_BLUE),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("BACKGROUND", (0, -1), (-1, -1), HexColor("#E2EFDA")),
            ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 4 * mm))


def _add_cost_summary_pdf(elements, styles, totals):
    """Page 6: Cost summary with a pie/donut chart placeholder."""
    elements.append(Paragraph("Cost Summary", styles["SectionHeading"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=PROPOSAL_BLUE,
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    # Donut chart placeholder — simulated with a styled table
    elements.append(Paragraph("<b>Cost Distribution</b>", styles["Normal"]))
    elements.append(Spacer(1, 3 * mm))

    cost_components = [
        ("Materials", totals.get("total_materials", 0)),
        ("Labour", totals.get("total_labour", 0)),
        ("Transport", totals.get("total_transport", 0)),
        ("Overhead", totals.get("total_overhead", 0)),
        ("Margin", totals.get("total_margin", 0)),
        ("GST", totals.get("total_gst", 0)),
    ]

    grand = totals.get("grand_total", 1) or 1
    chart_data = [["Component", "Amount", "% of Total"]]
    for label, amount in cost_components:
        pct = (amount / grand) * 100
        chart_data.append([
            label,
            f'{CURRENCY_SYMBOL}{amount:,.2f}',
            f"{pct:.1f}%",
        ])
    chart_data.append(["", "", ""])
    chart_data.append([
        "<b>Total</b>",
        f'<b>{CURRENCY_SYMBOL}{grand:,.2f}</b>',
        "<b>100%</b>",
    ])

    table = Table(chart_data, colWidths=[80 * mm, 50 * mm, 40 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), PROPOSAL_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("BACKGROUND", (0, -1), (-1, -1), HexColor("#E2EFDA")),
    ]))
    elements.append(table)

    # Text-based donut chart using Unicode
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph("<b>Distribution Visualization</b>", styles["Normal"]))
    elements.append(Spacer(1, 2 * mm))

    # Visual bar chart
    bar_chars = "▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓▓"
    colors_map = {
        "Materials": "🟦",
        "Labour": "🟩",
        "Transport": "🟨",
        "Overhead": "🟧",
        "Margin": "🟪",
        "GST": "🟥",
    }

    bar_data = []
    for label, amount in cost_components:
        pct = (amount / grand) * 100
        bar_len = max(1, int(pct / 5))
        bar = "█" * bar_len
        bar_data.append([
            f"{colors_map.get(label, '⬜')} {label}",
            f"{CURRENCY_SYMBOL}{amount:,.2f}",
            f"{pct:.1f}%",
        ])

    bar_table = Table(bar_data, colWidths=[100 * mm, 50 * mm, 30 * mm])
    bar_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, colors.lightgrey),
    ]))
    elements.append(bar_table)

    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph(
        "<i>* A visual pie/donut chart will be rendered here in the final version.</i>",
        ParagraphStyle(
            "Note",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER,
        ),
    ))


def _add_materials_highlights_pdf(elements, styles, item_dicts):
    """Page 7: Selected materials highlights."""
    elements.append(Paragraph("Selected Materials Highlights", styles["SectionHeading"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=PROPOSAL_BLUE,
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    # Filter items that have a material name
    materials_items = [
        d for d in item_dicts if d.get("material_name")
    ]

    if not materials_items:
        elements.append(Paragraph(
            "No specific materials have been selected for this project.",
            styles["Normal"],
        ))
        return

    mat_data = [["#", "Item Description", "Material", "Qty", "Unit", "Rate", "Total"]]
    for i, d in enumerate(materials_items[:20], 1):  # top 20
        qty = d.get("quantity", 0)
        rate = d.get("rate", 0)
        total = _r2(qty * rate)
        mat_data.append([
            str(i),
            (d.get("description", "") or "")[:50],
            d.get("material_name", "—"),
            f"{qty:.2f}",
            d.get("unit", "nos"),
            f'{CURRENCY_SYMBOL}{rate:,.2f}',
            f'{CURRENCY_SYMBOL}{total:,.2f}',
        ])

    col_widths = [8 * mm, 60 * mm, 40 * mm, 14 * mm, 10 * mm, 20 * mm, 22 * mm]
    table = Table(mat_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("BACKGROUND", (0, 0), (-1, 0), PROPOSAL_BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)


def _add_terms_and_conditions(elements, styles):
    """Page 8: Standard terms and conditions."""
    elements.append(Paragraph("Terms & Conditions", styles["SectionHeading"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=PROPOSAL_BLUE,
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    terms = [
        ("1. Scope of Work",
         "The cost estimate provided in this proposal covers the items and quantities "
         "listed in the Bill of Quantities. Any variations or additions to the scope "
         "shall be evaluated and quoted separately."),
        ("2. Validity",
         "This proposal is valid for a period of 30 days from the date of issue. "
         "Pricing is subject to change after the validity period."),
        ("3. Pricing & Currency",
         f"All prices are quoted in Indian Rupees ({CURRENCY_SYMBOL}) and include GST "
         "as applicable. The pricing is based on current market rates and supplier quotes."),
        ("4. Payment Terms",
         "Payment shall be made as per the agreed schedule: 30% advance on order "
         "acceptance, 40% on material delivery to site, 25% on completion of installation, "
         "and 5% retention for a period of 6 months."),
        ("5. Delivery & Installation",
         "Delivery timelines will be confirmed upon order placement. Installation "
         "schedules are dependent on site readiness and access."),
        ("6. Warranty",
         "All works carry a standard warranty of 12 months from the date of completion "
         "against manufacturing defects and workmanship issues."),
        ("7. Exclusions",
         "The following are excluded unless explicitly mentioned:\n"
         "• Structural changes to the building\n"
         "• Architectural / interior design fees\n"
         "• Statutory approvals and permits\n"
         "• Insurance beyond standard coverage\n"
         "• Force majeure conditions\n"
         "• GST / tax rate changes after the proposal date"),
        ("8. Acceptance",
         "This proposal shall be considered accepted upon receipt of a signed copy "
         "or purchase order referencing the proposal number."),
    ]

    for title, body in terms:
        elements.append(Paragraph(f"<b>{title}</b>", styles["TermsBody"]))
        elements.append(Paragraph(body, styles["TermsBody"]))
        elements.append(Spacer(1, 1 * mm))


# ======================================================================
# 3. Purchase List Excel
# ======================================================================


async def export_purchase_list(project_id: int, db) -> tuple[bytes, str]:
    """Generate a purchase list Excel grouped by vendor.

    Sheet 1: Items grouped by vendor
    Sheet 2: All items sorted by delivery urgency
    """
    project, boq_items, _latest_version = await _load_project_with_boq(project_id, db)
    item_dicts, breakdowns = _boq_items_to_dicts(boq_items)

    # Build vendor/item mapping from vendor_id
    vendor_items: dict[str, list[dict]] = {}
    for item_dict, bd in zip(item_dicts, breakdowns):
        vendor_id = item_dict.get("vendor_id")
        vendor_key = f"Vendor #{vendor_id}" if vendor_id else "Unassigned"
        vendor_items.setdefault(vendor_key, []).append({
            "description": item_dict.get("description", ""),
            "material_name": item_dict.get("material_name", "—"),
            "quantity": item_dict.get("quantity", 0),
            "unit": item_dict.get("unit", "nos"),
            "rate": item_dict.get("rate", 0),
            "total": _r2(bd.grand_total),
        })

    wb = Workbook()

    # ── Sheet 1: By Vendor ───────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Purchase by Vendor"
    ws1.sheet_properties.tabColor = "2E75B6"

    headers = ["Vendor", "Description", "Qty", "Unit", "Rate", "Amount"]
    col_widths = [22, 50, 12, 8, 14, 16]

    current_row = 1
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws1["A1"].value = f"Purchase List — {project.name}"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center")
    current_row = 3

    grand_total = 0.0
    for vendor_name in sorted(vendor_items.keys()):
        items = vendor_items[vendor_name]

        # Vendor header
        ws1.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(headers),
        )
        vh = ws1.cell(row=current_row, column=1, value=vendor_name)
        vh.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        vh.fill = SUBHEADER_FILL
        for c in range(1, len(headers) + 1):
            ws1.cell(row=current_row, column=c).fill = SUBHEADER_FILL
            ws1.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

        # Column headers
        for ci, h in enumerate(headers, 1):
            ws1.cell(row=current_row, column=ci, value=h)
        _style_header_row(ws1, current_row, len(headers))
        current_row += 1

        vendor_total = 0.0
        for item in items:
            total_val = item["total"]
            vendor_total += total_val
            row_data = ["", item["description"], item["quantity"], item["unit"], item["rate"], total_val]
            for ci, val in enumerate(row_data, 1):
                cell = ws1.cell(row=current_row, column=ci, value=val)
                _style_body_cell(cell, is_currency=(ci in (5, 6)))
            current_row += 1

        # Vendor subtotal
        ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        vs = ws1.cell(row=current_row, column=1, value=f"Subtotal — {vendor_name}")
        vs.font = TOTAL_FONT
        vs.fill = TOTAL_FILL
        vs.border = THIN_BORDER
        vs.alignment = Alignment(horizontal="right")
        for c in range(2, 5):
            ws1.cell(row=current_row, column=c).fill = TOTAL_FILL
            ws1.cell(row=current_row, column=c).border = THIN_BORDER
        amt_cell = ws1.cell(row=current_row, column=5, value=_r2(vendor_total))
        amt_cell.font = TOTAL_FONT
        amt_cell.fill = TOTAL_FILL
        amt_cell.border = THIN_BORDER
        amt_cell.number_format = INR_FORMAT
        ws1.cell(row=current_row, column=6).fill = TOTAL_FILL
        ws1.cell(row=current_row, column=6).border = THIN_BORDER
        current_row += 2
        grand_total += vendor_total

    # Grand total
    current_row += 1
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    gt_label = ws1.cell(row=current_row, column=1, value="GRAND TOTAL")
    gt_label.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    gt_label.fill = HEADER_FILL
    for c in range(2, 5):
        ws1.cell(row=current_row, column=c).fill = HEADER_FILL
        ws1.cell(row=current_row, column=c).border = THIN_BORDER
    gt_cell = ws1.cell(row=current_row, column=5, value=_r2(grand_total))
    gt_cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    gt_cell.fill = HEADER_FILL
    gt_cell.border = THIN_BORDER
    gt_cell.number_format = INR_FORMAT
    ws1.cell(row=current_row, column=6).fill = HEADER_FILL
    ws1.cell(row=current_row, column=6).border = THIN_BORDER

    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: By Urgency ──────────────────────────────────────────
    ws2 = wb.create_sheet("By Urgency")
    ws2.sheet_properties.tabColor = "C00000"

    urgency_headers = ["Item", "Description", "Qty", "Unit", "Rate", "Amount", "Vendor"]
    for ci, h in enumerate(urgency_headers, 1):
        ws2.cell(row=1, column=ci, value=h)
    _style_header_row(ws2, 1, len(urgency_headers))

    # Flatten all items with vendor info
    all_items_for_urgency: list[tuple[str, dict, str]] = []
    for vendor_name, items in vendor_items.items():
        for item in items:
            all_items_for_urgency.append((vendor_name, item, vendor_name))

    # Sort: unassigned first, then alphabetically by vendor
    all_items_for_urgency.sort(key=lambda x: (0 if x[0] == "Unassigned" else 1, x[0]))

    for i, (vendor_name, item, _) in enumerate(all_items_for_urgency, 1):
        row = 1 + i
        row_data = [
            i,
            item["description"],
            item["quantity"],
            item["unit"],
            item["rate"],
            item["total"],
            vendor_name,
        ]
        for ci, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=ci, value=val)
            _style_body_cell(cell, is_currency=(ci in (5, 6)))

    for ci, w in enumerate([5, 50, 12, 8, 14, 16, 22], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = project.name.replace(" ", "_").replace("/", "_")[:50] or f"project_{project_id}"
    filename = f"PurchaseList_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return buf.getvalue(), filename


# ======================================================================
# 4. Client Presentation PDF
# ======================================================================


async def export_client_presentation(project_id: int, db) -> tuple[bytes, str]:
    """Generate a client-facing presentation PDF.

    More marketing/visual focused than the proposal. Includes:
    - Cover with project branding
    - Project overview and scope
    - Material selections with placeholder images
    - Timeline / milestones (text-based Gantt)
    - Cost summary
    - Next steps
    """
    project, boq_items, latest_version = await _load_project_with_boq(project_id, db)
    item_dicts, breakdowns = _boq_items_to_dicts(boq_items)
    totals = _aggregate_python(breakdowns) if breakdowns else {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        title=f"Presentation — {project.name}",
        author="Auto Cost Engine",
    )

    styles = getSampleStyleSheet()
    _add_presentation_styles(styles)

    elements: list[Flowable] = []

    # ── Cover ─────────────────────────────────────────────────────────
    _add_presentation_cover(elements, styles, project)

    # ── Project Overview ──────────────────────────────────────────────
    elements.append(PageBreak())
    _add_presentation_overview(elements, styles, project, boq_items, totals)

    # ── Material Selections ───────────────────────────────────────────
    if item_dicts:
        elements.append(PageBreak())
        _add_presentation_materials(elements, styles, item_dicts)

    # ── Timeline / Milestones ─────────────────────────────────────────
    elements.append(PageBreak())
    _add_presentation_timeline(elements, styles, project)

    # ── Cost Summary ──────────────────────────────────────────────────
    elements.append(PageBreak())
    _add_presentation_cost_summary(elements, styles, totals)

    # ── Next Steps ────────────────────────────────────────────────────
    elements.append(PageBreak())
    _add_presentation_next_steps(elements, styles)

    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawCentredString(
            A4[0] / 2, 10 * mm,
            f"Confidential — {project.name} | Page {doc.page}",
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

    safe_name = project.name.replace(" ", "_").replace("/", "_")[:50] or f"project_{project_id}"
    filename = f"Presentation_{safe_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return buf.getvalue(), filename


def _add_presentation_styles(styles):
    """Add custom styles for the presentation."""
    accent_color = HexColor("#2E86C1")

    styles.add(
        ParagraphStyle(
            name="PresCoverTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=26,
            textColor=accent_color,
            spaceAfter=8 * mm,
            alignment=TA_CENTER,
        )
    )
    styles.add(
        ParagraphStyle(
            name="PresCoverSub",
            parent=styles["Normal"],
            fontSize=14,
            textColor=colors.HexColor("#555555"),
            spaceAfter=4 * mm,
            alignment=TA_CENTER,
        )
    )
    styles.add(
        ParagraphStyle(
            name="PresSection",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=16,
            textColor=accent_color,
            spaceBefore=4 * mm,
            spaceAfter=4 * mm,
        )
    )
    styles.add(
        ParagraphStyle(
            name="PresBody",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            spaceAfter=3 * mm,
        )
    )
    styles.add(
        ParagraphStyle(
            name="PresBullet",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            leftIndent=10 * mm,
            bulletIndent=0,
            spaceAfter=1 * mm,
        )
    )


def _add_presentation_cover(elements, styles, project):
    """Eye-catching cover page."""
    elements.append(Spacer(1, 30 * mm))

    # Accent bar
    elements.append(HRFlowable(
        width="100%", thickness=6, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=15 * mm,
    ))

    elements.append(Paragraph(project.name or "Project Proposal", styles["PresCoverTitle"]))
    if project.client:
        elements.append(Paragraph(
            f"Prepared for <b>{project.client}</b>", styles["PresCoverSub"],
        ))

    elements.append(Spacer(1, 10 * mm))

    info_lines = []
    if project.location:
        info_lines.append(f"📍 {project.location}")
    if project.project_code:
        info_lines.append(f"📋 Code: {project.project_code}")
    info_lines.append(f"📅 {datetime.now().strftime('%d %B %Y')}")

    for line in info_lines:
        elements.append(Paragraph(line, styles["PresCoverSub"]))

    elements.append(Spacer(1, 20 * mm))
    elements.append(HRFlowable(
        width="60%", thickness=2, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=0,
    ))

    elements.append(Spacer(1, 15 * mm))
    elements.append(Paragraph(
        "AUTO COST ENGINE",
        ParagraphStyle(
            "PresFooter",
            fontName="Helvetica-Bold",
            fontSize=10,
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER,
        ),
    ))


def _add_presentation_overview(elements, styles, project, boq_items, totals):
    """Project overview slide."""
    elements.append(Paragraph("Project Overview", styles["PresSection"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    if project.description:
        elements.append(Paragraph(project.description, styles["PresBody"]))
        elements.append(Spacer(1, 3 * mm))

    # Key metrics in a visual grid
    metrics = [
        ("Total Items", str(len(boq_items))),
        ("Total Area", "Computed from drawings"),
        ("Grand Total", f'{CURRENCY_SYMBOL}{totals.get("grand_total", 0):,.2f}'),
        ("Currency", project.currency),
    ]

    if project.start_date and project.end_date:
        duration = (project.end_date - project.start_date).days
        metrics.append(("Duration", f"{duration} days"))
        metrics.append(("Timeline", f"{project.start_date.strftime('%d-%b-%Y')} → {project.end_date.strftime('%d-%b-%Y')}"))

    metric_data = [["Metric", "Value"]]
    for label, value in metrics:
        metric_data.append([label, value])

    table = Table(metric_data, colWidths=[80 * mm, 80 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2E86C1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, HexColor("#EBF5FB")]),
    ]))
    elements.append(table)


def _add_presentation_materials(elements, styles, item_dicts):
    """Material selections with placeholder image boxes."""
    elements.append(Paragraph("Material Selections", styles["PresSection"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    # Filter items with material names, take top 6
    mat_items = [d for d in item_dicts if d.get("material_name")][:6]

    if not mat_items:
        elements.append(Paragraph(
            "No specific material selections have been made yet.",
            styles["PresBody"],
        ))
        return

    for i, item in enumerate(mat_items, 1):
        # Image placeholder box
        img_data = [
            ["",
             f"<b>{item.get('material_name', 'Material')}</b>",
             ""],
            ["",
             f"{item.get('description', '')[:40]}",
             ""],
            ["",
             f"Rate: {CURRENCY_SYMBOL}{item.get('rate', 0):,.2f}/{item.get('unit', 'nos')}",
             ""],
        ]
        img_table = Table(
            img_data,
            colWidths=[20 * mm, 120 * mm, 20 * mm],
        )
        img_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), HexColor("#F0F0F0")),
            ("BACKGROUND", (-1, 0), (-1, -1), HexColor("#F0F0F0")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(img_table)
        elements.append(Spacer(1, 2 * mm))

        # Alternate with description text
        elements.append(Paragraph(
            f"<b>{i}. {item.get('material_name', 'Material')}</b> — "
            f"{item.get('description', '')[:80]}",
            styles["PresBody"],
        ))


def _add_presentation_timeline(elements, styles, project):
    """Timeline / milestones slide using a text-based Gantt."""
    elements.append(Paragraph("Project Timeline & Milestones", styles["PresSection"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    # Build a text-based timeline
    if project.start_date and project.end_date:
        total_days = max(1, (project.end_date - project.start_date).days)

        milestones = [
            ("Project Initiation", 0, 10),
            ("Design & Planning", 10, 20),
            ("Procurement", 20, 35),
            ("Site Preparation", 30, 15),
            ("Main Installation", 45, 40),
            ("Finishing Works", 85, 25),
            ("Quality Inspection", 105, 10),
            ("Handover", 115, 10),
        ]

        # Scale down for A4
        chart_width = 150 * mm
        day_width = chart_width / total_days

        timeline_data = [["Phase", "Timeline"]]
        for label, start, duration in milestones:
            start_pct = (start / total_days) * 100
            dur_pct = (duration / total_days) * 100
            bar = "█" * max(1, int(dur_pct / 4))
            timeline_data.append([
                label,
                f"{' ' * int(start_pct / 6)}{bar}",
            ])

        table = Table(timeline_data, colWidths=[60 * mm, 100 * mm])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2E86C1")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, HexColor("#EBF5FB")]),
            ("FONTTEXT", (0, 1), (0, -1), "Helvetica"),
        ]))
        elements.append(table)

        elements.append(Spacer(1, 4 * mm))
        elements.append(Paragraph(
            f"<b>Start:</b> {project.start_date.strftime('%d-%b-%Y')} &nbsp;&nbsp; "
            f"<b>End:</b> {project.end_date.strftime('%d-%b-%Y')} &nbsp;&nbsp; "
            f"<b>Duration:</b> {total_days} days",
            styles["PresBody"],
        ))
    else:
        elements.append(Paragraph(
            "Timeline details are not yet configured for this project.",
            styles["PresBody"],
        ))

    # Milestone markers
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph("<b>Key Milestones</b>", styles["PresBody"]))
    milestones_text = [
        "✅ Project Kickoff — Day 0",
        "✅ Design Freeze — Day 15",
        "✅ Material Procurement Complete — Day 35",
        "✅ Installation Start — Day 45",
        "✅ Completion & Handover — Day 120+",
    ]
    for m in milestones_text:
        elements.append(Paragraph(f"• {m}", styles["PresBullet"]))


def _add_presentation_cost_summary(elements, styles, totals):
    """Cost summary slide."""
    elements.append(Paragraph("Cost Summary", styles["PresSection"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    cost_components = [
        ("Materials", totals.get("total_materials", 0)),
        ("Labour", totals.get("total_labour", 0)),
        ("Transport", totals.get("total_transport", 0)),
        ("Overhead", totals.get("total_overhead", 0)),
        ("Margin", totals.get("total_margin", 0)),
        ("GST", totals.get("total_gst", 0)),
    ]

    grand = totals.get("grand_total", 0)

    cost_data = [["Component", "Amount", "Share"]]
    for label, amount in cost_components:
        pct = (amount / grand * 100) if grand > 0 else 0
        cost_data.append([
            label,
            f'{CURRENCY_SYMBOL}{amount:,.2f}',
            f"{pct:.1f}%",
        ])
    cost_data.append(["", "", ""])
    cost_data.append([
        "<b>Grand Total</b>",
        f"<b>{CURRENCY_SYMBOL}{grand:,.2f}</b>",
        "<b>100%</b>",
    ])

    table = Table(cost_data, colWidths=[80 * mm, 50 * mm, 40 * mm])
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), HexColor("#2E86C1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("BACKGROUND", (0, -1), (-1, -1), HexColor("#D5F5E3")),
        ("LINEABOVE", (0, -1), (-1, -1), 1, HexColor("#1E8449")),
    ]))
    elements.append(table)

    # Visual bar chart
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph("<b>Cost Distribution</b>", styles["PresBody"]))

    bar_data = []
    for label, amount in cost_components:
        pct = (amount / grand * 100) if grand > 0 else 0
        bar_len = max(1, int(pct / 4))
        bar = "█" * min(bar_len, 30)
        bar_data.append([
            f"{label}",
            f'{CURRENCY_SYMBOL}{amount:,.2f}',
            f"{pct:.1f}%",
        ])

    bar_table = Table(bar_data, colWidths=[50 * mm, 50 * mm, 30 * mm])
    bar_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, colors.lightgrey),
    ]))
    elements.append(bar_table)


def _add_presentation_next_steps(elements, styles):
    """Next steps / call to action."""
    elements.append(Paragraph("Next Steps", styles["PresSection"]))
    elements.append(HRFlowable(
        width="100%", thickness=1, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=4 * mm,
    ))

    steps = [
        "1. <b>Review & Approve</b> — Please review the proposal and estimates at your convenience.",
        "2. <b>Scope Confirmation</b> — We will schedule a meeting to confirm the scope and address any queries.",
        "3. <b>Order Placement</b> — Upon approval, a formal purchase order will be issued.",
        "4. <b>Mobilization</b> — Project kickoff within 7 days of order confirmation.",
        "5. <b>Execution</b> — As per the agreed timeline and milestones.",
    ]

    for step in steps:
        elements.append(Paragraph(step, styles["PresBody"]))
        elements.append(Spacer(1, 2 * mm))

    elements.append(Spacer(1, 8 * mm))
    elements.append(HRFlowable(
        width="80%", thickness=1, color=HexColor("#2E86C1"),
        spaceBefore=0, spaceAfter=6 * mm,
    ))

    elements.append(Paragraph(
        "Thank you for considering Auto Cost Engine for your project needs.",
        ParagraphStyle(
            "ThankYou",
            fontName="Helvetica-Oblique",
            fontSize=11,
            textColor=colors.HexColor("#555555"),
            alignment=TA_CENTER,
        ),
    ))
    elements.append(Spacer(1, 4 * mm))
    elements.append(Paragraph(
        "<b>Contact:</b> hello@autocostengine.com | +91-XXXXXXXXXX",
        ParagraphStyle(
            "ContactLine",
            fontName="Helvetica",
            fontSize=9,
            textColor=colors.HexColor("#888888"),
            alignment=TA_CENTER,
        ),
    ))
